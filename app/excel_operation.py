import logging
from tkinter import messagebox, BooleanVar
import openpyxl
from app.database import Database
from datetime import datetime
import re


class ExcelOperation:
    """
    This class is responsible for processing the Excel file.
    """
    def __init__(self):
        self.folder = BooleanVar()
        self.file = BooleanVar()
        self.db = Database()
        self.error_messages = set()
        self.invalid_emails = []
        self.warning_messages = []

    def verify_format_file_excel(self, file_path):
        """
        This method verifies the format of the Excel file.
        Because the Excel file must be from app RelanceRH, it must contain the word "REPARTOIRE" in cell A1.
        """
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            repartoire = sheet['A1'].value
            if repartoire != "REPARTOIRE":
                self.error_messages.add(f"Email manquant ou format incorrect: {file_path}")
                return False
            return True
        except Exception as e:
            self.error_messages.add(f"Erreur lors de la vérification du format du fichier.")
            logging.error(f"Error verifying file format: {file_path}. Error: {e}")
            return False

    def process_excel_file(self, file_path):
        """
        This method processes the Excel file.
        """
        if self.verify_format_file_excel(file_path):
            info_from_excel = self.extract_information_from_excel(file_path)
            if info_from_excel:
                if self.invalid_emails:
                    self.warning_messages.append(
                        f"Les candidats souvent ont un email incorrect : {', '.join(self.invalid_emails)}"
                    )
                success = self.save_extracted_data(info_from_excel)
                self.show_warnings()
                return success
            else:
                self.error_messages.add(f"Erreur lors de l'extraction des informations du fichier: {file_path}, "
                                        f"verifiez le format des dates.")
                self.show_warnings()
                return False
        self.show_warnings()
        return False

    def extract_information_from_excel(self, file_path):
        """
        This method extracts the information from the Excel file.
        """
        try:
            wb = openpyxl.load_workbook(file_path)
            sheet = wb.active
            header_names = ['NOM', 'PRENOM', 'EMAIL', 'DERNIER ENTRETIEN']
            header_indices = {}

            for col_idx, cell in enumerate(sheet[1], start=1):
                if cell.value in header_names:
                    header_indices[cell.value] = col_idx - 1

            if len(header_indices) != len(header_names):
                self.error_messages.add(f"Le fichier ne contient pas tous les en-têtes nécessaires: {file_path}")
                logging.error(f"Missing headers in the file: {file_path}")
                raise ValueError("Not all expected headers found.")

            extracted_data = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                last_name = row[header_indices['NOM']]
                first_name = row[header_indices['PRENOM']]
                email_raw = row[header_indices['EMAIL']]
                last_interview_raw = row[header_indices['DERNIER ENTRETIEN']]
                last_interview = self.verify_format_date(last_interview_raw, f"{last_name} {first_name}", file_path)
                if last_interview is False:
                    continue

                if email_raw is None:
                    logging.error(f"Email missing in row: {row} of file: {file_path}")
                    continue


                email = self.clean_email_format(email_raw, f"{last_name} {first_name}")
                if email is None:
                    logging.error(f"Invalid email format in row: {row} of file: {file_path}")
                extracted_data.append({
                    'last_name': last_name,
                    'first_name': first_name,
                    'email': email,
                    'last_interview': last_interview
                })
            return extracted_data
        except Exception as e:
            self.error_messages.add(
                f"Problème lors de l'extraction des informations du fichier: {file_path}. Erreur: {e}")
            logging.error(f"Error extracting information from file: {file_path}. Error: {e}")
            return None

    def save_extracted_data(self, information):
        """
        This method saves the extracted data into the database.
        """
        success = False
        duplicates = []
        self.error_messages.clear()
        for info in information:
            try:
                data = {
                    'last_name': info['last_name'],
                    'first_name': info['first_name'],
                    'email': info['email'],
                    'last_interview': info['last_interview'],
                }
                if not self.verify_duplicate(data['email']):
                    self.db.insert_data(data)
                    success = True
                else:
                    duplicates.append(data['email'])
            except Exception as e:
                self.error_messages.add(f"Échec de l'enregistrement des données.")
                logging.error(f"Failed to save data: {info}. Error: {e}")
        if duplicates:
            self.warning_messages.append(
                f"Des doublons ont été trouvés pour les adresses suivantes: {', '.join(duplicates)}. Ces entrées seront ignorées."
            )

        return success

    def verify_duplicate(self, email):
        """
        This method verifies if the email already exists in the database.
        """
        data_bdd = self.db.get_user_by_email(email)
        if data_bdd:
            return True
        return False

    def verify_format_date(self, date, name, file_path):
        """
        This method verifies the format of the date.
        """
        if isinstance(date, datetime):
            return date.strftime('%d/%m/%y')
        if isinstance(date, str):
            match = re.match(r'(\d{1,2})/(\d{1,2})/(\d{2,4})', date)
            if match:
                day, month, year = match.groups()
                if len(year) == 2 or len(year) == 4:
                    corrected_date = f'{day}/{month}/{year}'
                    return corrected_date
                else:
                    logging.error(f"Invalid year format in date: {date} for {name}")
                    date = datetime.now().strftime('%d/%m/%y')
                    self.warning_messages.append(
                        f"Format de date incorrect pour {name}. On utilise la date actuelle."
                    )
                    return date
            else:
                logging.error(f"Invalid date format for {name}.")
                return False
        logging.error(f"Invalid date format for {name}.")
        return False

    def clean_email_format(self, email, name):
        """
        This method cleans the email format. we remove any spaces before or after the email, and we remove any spaces
        before or after the characters '@' and '.'
        """
        if email:
            # if "Mail:", ":", or "email" remove it
            email = re.sub(r'^(Mail\s*:\s*|:\s*|email\s*)', '', email, flags=re.IGNORECASE).strip()
            # Remove any spaces before or after the email
            email = re.sub(r'\s*([.@])\s*', r'\1', email)
            match = re.search(r'[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+', email)
            if match:
                email = match.group()
            else:
                logging.error(f"Invalid email format: {email}")
                self.invalid_emails.append(name)
                return None
        return email

    def show_warnings(self):
        """
        This method shows the warning messages.
        """
        if self.warning_messages:
            messagebox.showwarning("Avertissement", "\n".join(self.warning_messages))
        self.warning_messages.clear()